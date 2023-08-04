from openpyxl import load_workbook
import datetime
import os
import pandas as pd
# imports to email function
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# email function
def send_email(anexo, email):
    fn = pd.read_csv("Data/UserEmail.config", sep=";")
    # EMAIL AND PASSWORD
    EMAIL_ADRRESS = fn.email.to_string(index=False)  # Put your email in here
    EMAIL_PASSWORD = fn.senha.to_string(index=False)  # your email password, you can put a var whot open a file with your password
    EMAIL_RECIVER = email

    # CREATE EMAIL
    msg = MIMEMultipart()
    msg['Subject'] = 'Subject'
    msg['From'] = EMAIL_ADRRESS
    msg['To'] = EMAIL_RECIVER

    # Send
    body = "Programação impressa, Email gerado automaticamente"
    msg.attach(MIMEText(body, 'plain'))
    filename = anexo
    attachment = open(anexo, 'rb')
    p = MIMEBase('application', 'octet-stream')
    p.set_payload(attachment.read())
    encoders.encode_base64(p)

    # fixing the file
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)
    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(EMAIL_ADRRESS, EMAIL_PASSWORD)
    text = msg.as_string()
    s.sendmail(EMAIL_ADRRESS, EMAIL_RECIVER, text)
    s.quit()


# sheets in use
workbook = load_workbook('BaseInjecao.xlsx')
wb = load_workbook('BaseSopro.xlsx')
wbinjecao = workbook["Injeção"]
wbsopro = wb["Sopro"]

# Load ProductDatabase



class Injecao:
    def __init__(self,product_database ,machine_number, product, color, amount, bocal,client, color_cod, obs, vol):
        product_database = pd.read_csv('Data/product_database.csv', sep='|')
        self.name = None
        self.machine_number = machine_number.title()
        self.product = product
        self.color = color.title()
        self.amount = int(amount)
        self.bocal = bocal.title()
        self.client = client
        self.color_cod = color_cod
        self.obs = obs
        self.data = datetime.datetime.today().strftime('%d-%m-%y')
        self.vol = vol

        # Title
        wbinjecao['F4'] = f'Programação Da Máquina {self.machine_number}'

        # product

        wbinjecao['F9'] = self.product

        # cor
        wbinjecao['G9'] = self.color
        wbinjecao['N9'] = self.color
        
        #Cliente
        wbinjecao['M9'] = self.client

        # data
        wbinjecao['J6'] = f'Data: {self.data}'

        # that function get the amount will be produced, and return the total of boxes and the bulk in a box
        if_nan = product_database.query("Produto == '{}'".format(self.product)).volume.to_string(index=False)
        if if_nan == "NaN":
            wbinjecao['H9'] = f'{self.amount}'
            wbinjecao['O9'] = f'{self.amount}'
        else:
            verify_vol = int(product_database.query('Produto == "{}" and Setor == "Injecao"'.format(self.product)).get('volume'))
            wbinjecao['H9'] = f'{self.amount} ({int(self.amount / verify_vol)} {self.vol})'
            wbinjecao['O9'] = f'{verify_vol} (1 {self.vol})'

        # injecao_cli
        wbinjecao['K9'] = self.bocal

        # injecao_COD
        wbinjecao['P9'] = self.color_cod

        # injecao_OBs

        wbinjecao['Q9'] = self.obs

        # in here the name of file is created following Number of Machine - Date - Color
        # Save file

        self.name = f'M{self.machine_number}_{self.data}_{self.color}'
        workbook.save(f"Historico/{self.name}.xlsx")

    # the function find and get the file created to send email function in the script
    def send_email(self, email):
        anexo = f"Historico/{self.name}.xlsx"
        send_email(anexo, email)

    # it's the same but to impress
    def imprimir(self):
        anexo_nome = f'{self.name}.xlsx'
        file_path = f'Historico\\{anexo_nome}'
        os.startfile(file_path, 'print')


class Sopro:
    def __init__(self,product_database ,machine_number, product, color, amount, bocal ,client, obs):
        product_database = pd.read_csv('Data/product_database.csv', sep='|')
        self.name = None
        self.machine_number = machine_number.title()
        self.product = product
        self.color = color.title()
        self.amount = amount
        self.bocal = bocal
        self.client = client
        self.obs = obs
        self.data = datetime.datetime.today().strftime('%d-%m-%y')

        # Sopro Product
        wbsopro['B8'] = self.product
        

        # sopro_title
        s_title = f'Programação Da Máquina {self.machine_number} ({wbsopro["B8"].value}) Sopro'
        wbsopro['B3'] = s_title

        # sopro_color
        wbsopro['C8'] = self.color
        wbsopro['J8'] = self.color
        
        #client
        wbsopro['I8'] = self.client
        # Date

        wbsopro['F5'] = f'Data: {self.data}'

        # Sopro Amount
        wbsopro['D8'] = self.amount

        # bocal
        wbsopro['F8'] = self.bocal

        # sopro_OBs
        wbsopro['L8'] = self.obs

        # Save File
        self.name = f'Sopro M{self.machine_number} {self.data} {self.color}'
        wb.save(f"Historico/{self.name}.xlsx")

    def send_email(self, email):
        anexo = f"Historico/{self.name}.xlsx"
        send_email(anexo, email)

    def imprimir(self):
        anexo_nome = f'{self.name}.xlsx'
        file_path = f'Historico\\{anexo_nome}'
        os.startfile(file_path, 'print')

